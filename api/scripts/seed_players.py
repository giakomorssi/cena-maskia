"""Reset current team data and import teams + rosters from the Excel export.

Source workbook:
    c:\\Users\\FU249UR\\OneDrive - EY\\Desktop\\PERSONALE\\FANTA\\rose\\Rose_cena-maskia-championship.xlsx

What gets reset:
- teams
- players
- balances and entries
- transfers
- fines
- honors
- trade proposals and items

What is preserved:
- seasons
- news / polls / content

Imported team accounts are recreated as `squadra_1` .. `squadra_10` with password `utente`.

Salary rules from acquisition cost:
- 1..9    -> 0.5
- 10..19  -> 1.0
- 20..34  -> 2.0
- 35..49  -> 3.0
- 50..69  -> 4.5
- 70..89  -> 6.0
- 90+     -> 8.0
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import delete, select

from app.core.security import get_password_hash
from app.database import database_manager
from app.models.league import (
    BalanceEntry,
    BalanceSheet,
    Fine,
    Honor,
    Player,
    Season,
    Team,
    TradeProposal,
    TradeProposalItem,
    Transfer,
)
from app.services.player_finance_rules import fascia_from_cost, salary_from_cost

WORKBOOK_PATH = Path(
    r"c:\Users\FU249UR\OneDrive - EY\Desktop\PERSONALE\FANTA\rose\Rose_cena-maskia-championship.xlsx"
)


def _fascia_from_cost(cost: float) -> str:
    return fascia_from_cost(cost)


def _salary_from_cost(cost: float) -> float:
    return salary_from_cost(cost)


def _normalize_role(raw: Any) -> str:
    return str(raw or "").strip()


def _normalize_name(raw: Any) -> str:
    return str(raw or "").strip()


def _normalize_club(raw: Any) -> str:
    return str(raw or "").strip()


def _normalize_cost(raw: Any) -> float:
    value = str(raw or "0").strip().replace(",", ".")
    return float(value)


def _player_from_excel(role: Any, name: Any, club: Any, cost: Any) -> dict[str, Any]:
    amount = _normalize_cost(cost)
    return {
        "name": _normalize_name(name),
        "role": _normalize_role(role),
        "fascia": _fascia_from_cost(amount),
        "salary": _salary_from_cost(amount),
        "market_value": amount,
        "contract_years_total": 1,
        "contract_years_remaining": 1,
        "acquisition_type": "owned",
        "notes": f"Club origine: {_normalize_club(club)}" if club else None,
    }


def _iter_workbook_teams() -> list[tuple[str, list[dict[str, Any]]]]:
    if not WORKBOOK_PATH.exists():
        raise SystemExit(f"Workbook not found: {WORKBOOK_PATH}")

    ws = load_workbook(WORKBOOK_PATH, data_only=True).active
    teams: list[tuple[str, list[dict[str, Any]]]] = []

    for start_col in (1, 6):
        row = 5
        while row <= ws.max_row:
            team_name = ws.cell(row, start_col).value
            next_label = (
                ws.cell(row + 1, start_col).value if row + 1 <= ws.max_row else None
            )
            if team_name and str(next_label).strip() == "Ruolo":
                players: list[dict[str, Any]] = []
                current_team = str(team_name).strip()
                row += 2
                while row <= ws.max_row:
                    role = ws.cell(row, start_col).value
                    name = ws.cell(row, start_col + 1).value
                    club = ws.cell(row, start_col + 2).value
                    cost = ws.cell(row, start_col + 3).value

                    role_text = str(role or "").strip()
                    if role_text.startswith("Crediti Residui:"):
                        break
                    if all(v is None for v in (role, name, club, cost)):
                        break
                    if name:
                        players.append(_player_from_excel(role, name, club, cost))
                    row += 1

                if players:
                    teams.append((current_team, players))
            row += 1

    return teams


def _reset_team_linked_data(db) -> None:
    db.execute(delete(TradeProposalItem))
    db.execute(delete(TradeProposal))
    db.execute(delete(Transfer))
    db.execute(delete(BalanceEntry))
    db.execute(delete(BalanceSheet))
    db.execute(delete(Fine))
    db.execute(delete(Honor))
    db.execute(delete(Player))
    db.execute(delete(Team))
    db.flush()


def main() -> None:
    try:
        database_manager.initialize()
    except Exception:
        pass

    imported_teams = _iter_workbook_teams()
    if len(imported_teams) != 10:
        raise SystemExit(
            f"Expected 10 non-empty teams from workbook, found {len(imported_teams)}"
        )

    with database_manager.get_session() as db:
        season = (
            db.execute(select(Season).where(Season.is_current.is_(True)))
            .scalars()
            .first()
        )
        if season is None:
            season = (
                db.execute(select(Season).order_by(Season.year.desc()))
                .scalars()
                .first()
            )
        if season is None:
            raise SystemExit("No season found. Create a season first.")

        print(f"Using season: {season.name} (id={season.id})")
        print("Resetting current team-linked data...")
        _reset_team_linked_data(db)

        team_count = 0
        player_count = 0
        for index, (team_name, players) in enumerate(imported_teams, start=1):
            team = Team(
                name=team_name,
                account_username=f"squadra_{index}",
                password_hash=get_password_hash("utente"),
                manager_name=team_name,
                profile_bio="Importato da Rose_cena-maskia-championship.xlsx",
                is_active=True,
            )
            db.add(team)
            db.flush()
            team_count += 1
            print(f"-> Team {index}: {team_name} ({len(players)} giocatori)")

            for payload in players:
                db.add(
                    Player(
                        team_id=team.id,
                        season_id=season.id,
                        acquisition_season_id=season.id,
                        is_active=True,
                        **payload,
                    )
                )
                player_count += 1

        db.commit()
        print(f"\nDone. Teams={team_count}, Players={player_count}")


if __name__ == "__main__":
    main()
