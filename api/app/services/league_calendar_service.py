from __future__ import annotations

from dataclasses import dataclass
from typing import TYPE_CHECKING

from openpyxl import load_workbook

if TYPE_CHECKING:
    from sqlalchemy.orm import Session


@dataclass(frozen=True)
class ScheduleMatch:
    home_team: str
    home_score: float | None
    away_score: float | None
    away_team: str
    result: str | None


@dataclass(frozen=True)
class ScheduleRound:
    league_round: int
    serie_a_round: int | None
    matches: list[ScheduleMatch]


@dataclass(frozen=True)
class StandingRow:
    position: int
    team: str
    played: int
    wins: int
    draws: int
    losses: int
    goals_for: int
    goals_against: int
    goal_diff: int
    points: int
    total_points: float


@dataclass(frozen=True)
class LeagueCalendarData:
    title: str
    source_url: str | None
    standings: list[StandingRow]
    rounds: list[ScheduleRound]


# ---------------------------------------------------------------------------
# DB-backed read (primary)
# ---------------------------------------------------------------------------


def get_league_calendar_data_from_db(db: "Session") -> LeagueCalendarData:
    """Read standings and calendar from the database."""
    from sqlalchemy import select
    from app.models.league import CalendarRound, Season
    from app.models.league import StandingRow as StandingRowModel

    current_season = db.execute(
        select(Season).where(Season.is_current.is_(True))
    ).scalar_one_or_none()

    standings: list[StandingRow] = []
    rounds: list[ScheduleRound] = []
    title = "Calendario"

    if current_season:
        title = f"Stagione {current_season.name}"

        db_standings = (
            db.execute(
                select(StandingRowModel)
                .where(StandingRowModel.season_id == current_season.id)
                .order_by(StandingRowModel.position)
            )
            .scalars()
            .all()
        )

        standings = [
            StandingRow(
                position=r.position,
                team=r.team_name,
                played=r.played,
                wins=r.wins,
                draws=r.draws,
                losses=r.losses,
                goals_for=r.goals_for,
                goals_against=r.goals_against,
                goal_diff=r.goal_diff,
                points=r.points,
                total_points=r.total_points,
            )
            for r in db_standings
        ]

        db_rounds = (
            db.execute(
                select(CalendarRound)
                .where(CalendarRound.season_id == current_season.id)
                .order_by(CalendarRound.league_round)
            )
            .scalars()
            .all()
        )

        for rnd in db_rounds:
            matches = [
                ScheduleMatch(
                    home_team=m.home_team,
                    home_score=m.home_score,
                    away_score=m.away_score,
                    away_team=m.away_team,
                    result=m.result,
                )
                for m in sorted(rnd.matches, key=lambda x: x.match_order)
            ]
            rounds.append(
                ScheduleRound(
                    league_round=rnd.league_round,
                    serie_a_round=rnd.serie_a_round,
                    matches=matches,
                )
            )

    return LeagueCalendarData(
        title=title, source_url=None, standings=standings, rounds=rounds
    )


# ---------------------------------------------------------------------------
# Excel parsing helpers (used by upload endpoint to populate DB)
# ---------------------------------------------------------------------------


def _parse_round_number(label: str) -> int:
    digits = "".join(char for char in label if char.isdigit())
    if not digits:
        raise ValueError(f"Unable to parse round number from {label!r}")
    return int(digits)


def _clean_text(value: object) -> str:
    return str(value).replace("â°", "°").replace("�", "°").strip()


def _load_sheet(content: bytes):
    from io import BytesIO

    return load_workbook(BytesIO(content), data_only=True).active


def parse_standings_excel(content: bytes) -> list[dict]:
    """Parse standings Excel bytes → list of row dicts."""
    ws = _load_sheet(content)
    rows = []
    for row in range(5, ws.max_row + 1):
        position = ws.cell(row=row, column=1).value
        team = ws.cell(row=row, column=2).value
        if position is None or not team:
            continue
        rows.append(
            {
                "position": int(position),
                "team_name": _clean_text(team),
                "played": int(ws.cell(row=row, column=4).value or 0),
                "wins": int(ws.cell(row=row, column=5).value or 0),
                "draws": int(ws.cell(row=row, column=6).value or 0),
                "losses": int(ws.cell(row=row, column=7).value or 0),
                "goals_for": int(ws.cell(row=row, column=8).value or 0),
                "goals_against": int(ws.cell(row=row, column=9).value or 0),
                "goal_diff": int(ws.cell(row=row, column=10).value or 0),
                "points": int(ws.cell(row=row, column=11).value or 0),
                "total_points": float(ws.cell(row=row, column=12).value or 0),
            }
        )
    return rows


def parse_calendar_excel(content: bytes) -> list[dict]:
    """Parse calendar Excel bytes → list of round dicts with nested matches."""
    ws = _load_sheet(content)
    rounds = []
    row = 4
    while row <= ws.max_row:
        left_header = ws.cell(row=row, column=1).value
        right_header = ws.cell(row=row, column=7).value
        if left_header is None and right_header is None:
            row += 1
            continue

        for start_col, header_value, serie_a_col in (
            (1, left_header, 3),
            (7, right_header, 9),
        ):
            if not header_value:
                continue
            league_round = _parse_round_number(_clean_text(header_value))
            serie_a_value = ws.cell(row=row, column=serie_a_col).value
            serie_a_round = (
                _parse_round_number(_clean_text(serie_a_value))
                if serie_a_value
                else None
            )
            matches = []
            for i, match_row in enumerate(range(row + 1, min(row + 6, ws.max_row + 1))):
                home_team = ws.cell(row=match_row, column=start_col).value
                away_team = ws.cell(row=match_row, column=start_col + 3).value
                result = ws.cell(row=match_row, column=start_col + 4).value
                if not home_team and not away_team:
                    continue
                home_score_val = ws.cell(row=match_row, column=start_col + 1).value
                away_score_val = ws.cell(row=match_row, column=start_col + 2).value
                matches.append(
                    {
                        "match_order": i,
                        "home_team": _clean_text(home_team or "-"),
                        "away_team": _clean_text(away_team or "-"),
                        "home_score": (
                            float(home_score_val)
                            if home_score_val is not None
                            else None
                        ),
                        "away_score": (
                            float(away_score_val)
                            if away_score_val is not None
                            else None
                        ),
                        "result": _clean_text(result) if result else None,
                    }
                )
            rounds.append(
                {
                    "league_round": league_round,
                    "serie_a_round": serie_a_round,
                    "matches": matches,
                }
            )

        row += 6

    return rounds


def parse_rose_excel(content: bytes) -> list[dict]:
    """Parse rose Excel bytes → list of {team_name, players} dicts."""
    from app.services.player_finance_rules import fascia_from_cost, salary_from_cost

    ws = _load_sheet(content)
    teams = []

    for start_col in (1, 6):
        row = 5
        while row <= ws.max_row:
            team_name = ws.cell(row, start_col).value
            next_label = (
                ws.cell(row + 1, start_col).value if row + 1 <= ws.max_row else None
            )
            if team_name and str(next_label).strip() == "Ruolo":
                players = []
                current_team = str(team_name).strip()
                row += 2
                while row <= ws.max_row:
                    role = ws.cell(row, start_col).value
                    name = ws.cell(row, start_col + 1).value
                    club = ws.cell(row, start_col + 2).value
                    cost_raw = ws.cell(row, start_col + 3).value
                    if role is None and name is None:
                        break
                    if name:
                        cost = float(
                            str(cost_raw or "0").strip().replace(",", ".") or 0
                        )
                        players.append(
                            {
                                "name": str(name).strip(),
                                "role": str(role or "").strip(),
                                "fascia": fascia_from_cost(cost),
                                "salary": salary_from_cost(cost),
                                "market_value": cost,
                                "contract_years_total": 1,
                                "contract_years_remaining": 1,
                                "acquisition_type": "owned",
                                "notes": (
                                    f"Club origine: {str(club).strip()}"
                                    if club
                                    else None
                                ),
                                "is_active": True,
                            }
                        )
                    row += 1
                teams.append({"team_name": current_team, "players": players})
            else:
                row += 1

    return teams
